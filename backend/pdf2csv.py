import pdftables_api
import os

from openpyxl import load_workbook
import pandas as pd


def convert_to_csv(input, output):
    API_KEY = os.environ['PDFTABLES_API_KEY']
    c = pdftables_api.Client(API_KEY)
    try:
        c.xlsx(input, output)
        c.csv(input, output)
    except Exception as e:
        print(e)


def normalize_data_rs(input, output):
    wb = load_workbook(input)
    ws = wb['Sheet1']
    colE = ws['E']

    for cell in colE:
        nama_rs = cell.value
        new_cell = ws.cell(row=cell.row, column=4)
        if 'RSUD ' in nama_rs:
            new_cell.value = 'Rumah Sakit Umum Daerah'
        elif 'RS Jiwa ' in nama_rs:
            new_cell.value = 'Rumah Sakit Jiwa'
        elif ('RSU ' in nama_rs) or ('RS ' in nama_rs):
            new_cell.value = 'Rumah Sakit Umum'
        elif 'RSK ' in nama_rs:
            new_cell.value = 'Rumah Sakit Khusus'
    wb.save(output + '.xlsx')
